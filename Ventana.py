import tkinter as tk
import tkinter.filedialog
from tkinter import messagebox

import openpyxl
from openpyxl import load_workbook, Workbook

#importar librerias: pip openpyxl...

hoja = None
origen = None

def abrir_archivo_origen():
    global origen, hoja
    file = tkinter.filedialog.askopenfilename()
    origen = load_workbook(file)
    #Hoja1 debido a que el excel esta en este caso configurado en español, sino seria sheet
    hoja = origen['Hoja1']

def seleccionar_destino():
    global destino, destino_ruta
    destino_ruta = tkinter.filedialog.asksaveasfilename(filetypes=[("Excel files", "*.xlsx")])
    destino = destino_ruta + ".xlsx"

def leer_datos():
    dato_1.insert(0, hoja['A1'].value)
    dato_2.insert(0, hoja['A2'].value)
    dato_3.insert(0, hoja['A3'].value)
    dato_4.insert(0, hoja['A4'].value)
    dato_5.insert(0, hoja['A5'].value)


def editar_datos():
    global nuevo_dato_1, nuevo_dato_2, nuevo_dato_3, nuevo_dato_4, nuevo_dato_5, valores
    nuevo_dato_1 = dato_1.get()
    nuevo_dato_2 = dato_2.get()
    nuevo_dato_3 = dato_3.get()
    nuevo_dato_4 = dato_4.get()
    nuevo_dato_5 = dato_5.get()

    valores = [dato_1.get(), dato_2.get(), dato_3.get(), dato_4.get(), dato_5.get()]

def guardar_datos(ruta_absoluta, valores):
    excel_nuevo = openpyxl.Workbook()
    hoja = excel_nuevo.active

    for i, valor in enumerate(valores, 1):
        hoja[f"A{i}"] = valor

    excel_nuevo.save(ruta_absoluta)

    messagebox.showinfo("Nuevo Excel Guardado", "Nuevo excel guardado en la ruta deseada.")
    root.destroy()


root = tk.Tk()
root.geometry("400x300")

boton_selecionar_origen = tk.Button(root, text="Seleccionar Origen", command=abrir_archivo_origen)
boton_leer_datos = tk.Button(root, text="Leer Datos", command=leer_datos)
boton_editar_datos = tk.Button(root, text="Editar Datos", command=editar_datos)
boton_selecionar_destino = tk.Button(root, text="Seleccionar Destino", command=seleccionar_destino)
dato_1 = tk.Entry(root, width=10)
dato_2 = tk.Entry(root, width=10)
dato_3 = tk.Entry(root, width=10)
dato_4 = tk.Entry(root, width=10)
dato_5 = tk.Entry(root, width=10)


boton_escribir_datos = tk.Button(root, text="Escribir Datos", command=lambda: guardar_datos(destino, valores))
boton_selecionar_origen.grid(row=1, column=1)
boton_leer_datos.grid(row=2, column=1)
boton_editar_datos.grid(row=6, column=3)
boton_selecionar_destino.grid(row=1, column=5)
boton_escribir_datos.grid(row=2, column=5)
dato_1.grid(row=1, column=2)
dato_2.grid(row=2, column=2)
dato_3.grid(row=3, column=2)
dato_4.grid(row=4, column=2)
dato_5.grid(row=5, column=2)

root.mainloop()
